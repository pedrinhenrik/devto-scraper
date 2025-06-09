def gerar_relatorio_excel():
    import pandas as pd
    import mysql.connector
    from openpyxl import load_workbook
    from openpyxl.drawing.image import Image as Img
    from openpyxl.styles import Alignment, Font, PatternFill
    import matplotlib.pyplot as plt
    import numpy as np
    import os
    from wordcloud import WordCloud
    from src.config import DB_HOST, DB_USER, DB_PASSWORD, DB_NAME
    from collections import Counter
    from openpyxl.drawing.image import Image

    # Cria o caminho da pasta
    pasta_relatorios = os.path.join(os.path.dirname(__file__), '..', 'relatorios')
    os.makedirs(pasta_relatorios, exist_ok=True)
    arquivo_excel = os.path.join(pasta_relatorios, 'relatorio_posts_devto.xlsx')

    # Conexão ao banco
    conn = mysql.connector.connect(
        host=DB_HOST,
        user=DB_USER,
        password=DB_PASSWORD,
        database=DB_NAME
    )

    # Query para pegar todos os posts
    query = "SELECT * FROM posts"
    df = pd.read_sql(query, conn)

    # Exporta para Excel
    df.to_excel(arquivo_excel, index=False)

    # Abre o arquivo para formatação
    wb = load_workbook(arquivo_excel)
    ws = wb.active
    ws.title = "Posts"
    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = ws.dimensions

    # Centralizar todas as células da aba Posts
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(horizontal='center', vertical='center')

    # Formatar o cabeçalho
    header_fill = PatternFill(start_color='006400', end_color="224e22", fill_type='solid')
    header_font = Font(color='FFFFFF', bold=True)

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font

    # Ajustar largura das colunas
    for column_cells in ws.columns:
        max_length = 0
        column = column_cells[0].column_letter
        for cell in column_cells:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = adjusted_width

    def encurtar_texto(texto, limite=40):
        if len(texto) > limite:
            return texto[:limite] + '...'
        return texto

    # Gráfico Top 10 autores
    top_autores = df['autor'].value_counts().head(10)
    fig, ax = plt.subplots(figsize=(10, 6))
    colors = ['#00bfff' if val < top_autores.max() else '#ff4500' for val in top_autores]
    bars = ax.barh(top_autores.index, top_autores.values, color=colors, edgecolor='black')
    fig.suptitle('Top 10 Autores por Número de Posts', fontsize=16, weight='bold', y=0.98)
    ax.set_xlabel('Quantidade de Posts', fontsize=12)
    ax.set_ylabel('Autores', fontsize=12)

    for bar in bars:
        width = bar.get_width()
        ax.annotate(f'{width}', xy=(width, bar.get_y() + bar.get_height() / 2),
                    xytext=(5, 0), textcoords='offset points',
                    ha='left', va='center', fontsize=10, fontweight='bold', color='black')
    for spine in ax.spines.values():
        spine.set_visible(False)
    plt.tight_layout()
    caminho_img_autores = os.path.join(pasta_relatorios, 'top_autores.png')
    plt.savefig(caminho_img_autores)
    plt.close()

    # Junta todas as tags
    todas_tags = []
    for tags in df['tags']:
        if pd.notnull(tags):
            tags_list = [tag.strip() for tag in tags.split(',')]
            todas_tags.extend(tags_list)

    contador_tags = Counter(todas_tags)
    top_tags = contador_tags.most_common(10)
    tags_labels, tags_counts = zip(*top_tags)
    fig, ax = plt.subplots(figsize=(10, 6))
    colors = ['#00bfff' if val < max(tags_counts) else '#ff4500' for val in tags_counts]
    bars = ax.barh(tags_labels, tags_counts, color=colors, edgecolor='black')
    fig.suptitle('Top 10 Tags Mais Usadas', fontsize=16, weight='bold', y=0.98)
    ax.set_xlabel('Quantidade de Vezes', fontsize=12)
    ax.set_ylabel('Tag', fontsize=12)

    for bar in bars:
        width = bar.get_width()
        ax.annotate(f'{width}', xy=(width, bar.get_y() + bar.get_height() / 2),
                    xytext=(5, 0), textcoords='offset points',
                    ha='left', va='center', fontsize=10, fontweight='bold', color='black')
    for spine in ax.spines.values():
        spine.set_visible(False)
    plt.tight_layout()
    caminho_img_tags = os.path.join(pasta_relatorios, 'top_tags.png')
    plt.savefig(caminho_img_tags)
    plt.close()

    # Gráfico Top 10 posts por reações
    top_reacoes = df.sort_values(by='reacoes', ascending=False).head(10)
    top_reacoes['titulo_curto'] = top_reacoes['titulo'].apply(lambda x: encurtar_texto(str(x), limite=20))
    fig, ax = plt.subplots(figsize=(10, 6))
    colors = ['#00bfff' if val < top_reacoes['reacoes'].max() else '#ff4500' for val in top_reacoes['reacoes']]
    bars = ax.barh(top_reacoes['titulo_curto'], top_reacoes['reacoes'], color=colors, edgecolor='black')
    fig.suptitle('Top 10 Posts por Número de Reações', fontsize=16, weight='bold', y=0.98)
    ax.set_xlabel('Reações', fontsize=12)
    ax.set_ylabel('Título', fontsize=12)

    for bar in bars:
        width = bar.get_width()
        ax.annotate(f'{width}', xy=(width, bar.get_y() + bar.get_height() / 2),
                    xytext=(5, 0), textcoords='offset points',
                    ha='left', va='center', fontsize=10, fontweight='bold', color='black')
    for spine in ax.spines.values():
        spine.set_visible(False)
    plt.tight_layout()
    caminho_img_reacoes = os.path.join(pasta_relatorios, 'top_posts_reacoes.png')
    plt.savefig(caminho_img_reacoes)
    plt.close()

    # Gráfico de distribuição de reações e comentários
    bins = [0, 5, 10, 20, 30, 40, 50, float('inf')]
    labels = ['0-2', '3-5', '6-10','11-20', '21-30', '31-50', '51+']
    faixas_reacoes = pd.cut(df['reacoes'], bins=bins, labels=labels, right=True)
    faixas_comentarios = pd.cut(df['comentarios'], bins=bins, labels=labels, right=True)
    contagem_reacoes = faixas_reacoes.value_counts().sort_index()
    contagem_comentarios = faixas_comentarios.value_counts().sort_index()

    x = np.arange(len(labels))
    width = 0.35

    # Criar gráfico de barras
    fig, ax = plt.subplots(figsize=(10, 6))
    bars1 = ax.bar(x - width/2, contagem_reacoes, width, label='Reações', color='#00bfff', edgecolor='black')
    bars2 = ax.bar(x + width/2, contagem_comentarios, width, label='Comentários', color='#ff4500', edgecolor='black')

    # Título e labels
    fig.suptitle('Distribuição de Faixas de Reações e Comentários', fontsize=16, weight='bold', y=0.98)
    ax.set_xlabel('Faixas', fontsize=12)
    ax.set_ylabel('Quantidade de Posts', fontsize=12)

    ax.set_xticks(x)
    ax.set_xticklabels(labels, rotation=45)
    ax.legend()

    # Adicionando os valores no topo das barras
    for bar in bars1 + bars2:
        height = bar.get_height()
        if height > 0:
            ax.annotate(f'{int(height)}',
                        xy=(bar.get_x() + bar.get_width() / 2, height),
                        xytext=(0, 3),
                        textcoords="offset points",
                        ha='center', va='bottom', fontsize=10, fontweight='bold', color='black')

    # Remover bordas
    for spine in ax.spines.values():
        spine.set_visible(False)

    def gerar_wordcloud_tags(df, pasta_relatorios):
        todas_tags = []
        for tags in df['tags']:
            if pd.notnull(tags):
                tags_list = [tag.strip().replace('#', '') for tag in tags.split(',')]
                todas_tags.extend(tags_list)

        # Junta tudo em uma única string separada por espaços
        texto_tags = ' '.join(todas_tags)

        # Gera a WordCloud
        wordcloud = WordCloud(
            width=800, 
            height=400, 
            background_color='white', 
            colormap='plasma',
            max_font_size=120,
            random_state=42
        ).generate(texto_tags)

        # Salva a imagem
        caminho_img_wordcloud = os.path.join(pasta_relatorios, 'wordcloud_tags.png')
        plt.figure(figsize=(12, 6))
        plt.imshow(wordcloud, interpolation='bilinear')
        plt.axis('off')
        plt.tight_layout(pad=0)
        plt.savefig(caminho_img_wordcloud)
        plt.close()

        return caminho_img_wordcloud

    plt.tight_layout()
    caminho_img_distribuicao = os.path.join(pasta_relatorios, 'distribuicao_reacoes_comentarios.png')
    plt.savefig(caminho_img_distribuicao)
    plt.close()

    # Criar abas e adicionar imagens
    aba_autores = wb.create_sheet(title="Gráfico Autores")
    img_autores = Img(caminho_img_autores)
    img_autores.anchor = 'A1'
    aba_autores.add_image(img_autores)

    aba_reacoes = wb.create_sheet(title="Gráfico Reações")
    img_reacoes = Img(caminho_img_reacoes)
    img_reacoes.anchor = 'A1'
    aba_reacoes.add_image(img_reacoes)

    aba_tags = wb.create_sheet(title="Gráfico Tags")
    img_tags = Img(caminho_img_tags)
    img_tags.anchor = 'A1'
    aba_tags.add_image(img_tags)

    aba_distribuicao = wb.create_sheet(title="Gráfico Distribuição")
    img_distribuicao = Img(caminho_img_distribuicao)
    img_distribuicao.anchor = 'A1'
    aba_distribuicao.add_image(img_distribuicao)

    # Gera a WordCloud e salva
    caminho_img_wordcloud = gerar_wordcloud_tags(df, pasta_relatorios)

    # Cria nova aba no Excel e adiciona a WordCloud
    aba_wordcloud = wb.create_sheet(title="Wordcloud Tags")
    img_wordcloud = Image(caminho_img_wordcloud)
    img_wordcloud.anchor = 'A1'
    aba_wordcloud.add_image(img_wordcloud)


    wb.save(arquivo_excel)
    
    print(f"✅ Relatório gerado: {arquivo_excel}")
